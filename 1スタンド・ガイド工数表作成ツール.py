import streamlit as st
import pandas as pd
import openpyxl
import io
import os

st.title("スタンド・ガイド正規出図工数表作成ツール")

options = [chr(i) for i in range(ord('A'), ord('L') + 1)] + ["（空白）"]
selected_letter = st.selectbox("出荷区分を選択", options)

company_options = [
    "PAP",
    "Y・Gテック",
    "ヒラテ技研（近江八幡メンバー）",
    "ヒラテ技研（請負メンバー）",
    "DSE",
    "ユニテツク",
    "タイガ設計",
    "中央エンジ"
]

stand_company = st.selectbox("スタンド設計会社を選択", company_options)
guide_company = st.selectbox("ガイド設計会社を選択", company_options)

uploaded_file = st.file_uploader("📁 仕様一覧表(xlsm)をアップロードしてください", type=["xlsx", "xlsm"])

# テンプレートファイル名を変更
template_path = "template.xlsx"

if not os.path.exists(template_path):
    st.error("❌ テンプレートファイルが見つかりません。GitHub に 'template.xlsx' を正しくアップロードしてください。")
elif uploaded_file:
    try:
        uploaded_bytes = uploaded_file.read()
        uploaded_io = io.BytesIO(uploaded_bytes)
        df = pd.read_excel(uploaded_io, sheet_name="仕様一覧表", engine="openpyxl")

        expected_n_idx = 13
        expected_ac_idx = 28
        if expected_n_idx >= len(df.columns) or expected_ac_idx >= len(df.columns):
            st.error("❌ 仕様一覧表に想定した列が存在しません。列数を確認してください。")
        else:
            n_col = df.columns[expected_n_idx]
            ac_col = df.columns[expected_ac_idx]

            if selected_letter == "（空白）":
                filtered_df = df[df[n_col].isna() | (df[n_col].astype(str).str.strip() == "")]
            else:
                filtered_df = df[df[n_col] == selected_letter]

            ac_series = filtered_df[ac_col].astype(str).str.replace(r"[^\d.]", "", regex=True)
            ac_series = ac_series.replace("", pd.NA).dropna()
            ac_cleaned = pd.to_numeric(ac_series, errors="coerce").dropna()

            total_ac = ac_cleaned.sum() if not ac_cleaned.empty else 0.0
            total_distance = total_ac / 1000.0
            count = len(filtered_df)

            if count == 0:
                st.warning(f"⚠️ N列に '{selected_letter}' が含まれる行は見つかりませんでした。")
            else:
                label = "空白" if selected_letter == "（空白）" else selected_letter
                st.success(f"✅ 出荷区分が '{label}' （{count} 機番）の総機長は：{total_distance:.2f}m")
                st.dataframe(filtered_df)

                wb = openpyxl.load_workbook(template_path)
                target_sheets = {
                    "スタンド正規出図": stand_company,
                    "ガイド正規出図": guide_company
                }

                for sheet_name, company_value in target_sheets.items():
                    if sheet_name not in wb.sheetnames:
                        st.error(f"❌ テンプレートにシート '{sheet_name}' が存在しません。")
                        continue
                    ws = wb[sheet_name]
                    found_total = False
                    found_company = False
                    for row in ws.iter_rows(values_only=False):
                        for cell in row:
                            if cell.value == "総距離(m)":
                                ws.cell(row=cell.row + 1, column=cell.column).value = total_distance
                                found_total = True
                            if cell.value == "計算式":
                                ws.cell(row=cell.row - 1, column=cell.column).value = company_value
                                found_company = True
                    if not found_total:
                        st.warning(f"☝️ シート '{sheet_name}' に '総距離(m)' が見つかりませんでした。")
                    if not found_company:
                        st.warning(f"☝️ シート '{sheet_name}' に '計算式' が見つかりませんでした。")

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="📥 スタンド・ガイド工数表をダウンロード",
                    data=output.getvalue(),
                    file_name="スタンドガイド工数計画_更新済.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"❌ エラーが発生しました：{str(e)}")
else:
    st.info("👆 上のボックスにExcelファイルをドラッグ＆ドロップしてください")
